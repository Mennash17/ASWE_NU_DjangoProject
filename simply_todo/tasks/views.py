from django.shortcuts import render, redirect, get_object_or_404
from .models import Task
from .forms import TaskForm
from django.db.models import Q
from django.http import JsonResponse
from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.forms import UserCreationForm
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .serializers import TaskSerializer
from rest_framework import generics, permissions
from rest_framework.authentication import TokenAuthentication
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
import openpyxl

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm
from django.http import JsonResponse


from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login
from django.http import HttpResponseRedirect

from rest_framework.views import APIView
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from rest_framework.response import Response
from rest_framework import status

def custom_login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            else:
                return redirect('task_list')
    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def export_tasks_excel_api(request):
    user = request.user
    tasks = Task.objects.filter(user=user)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'My Tasks'

    headers = ['Title', 'Description', 'Category', 'Due Date', 'Priority', 'Completed', 'Notes']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

    for col_num, column_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_num, task in enumerate(tasks, start=2):
        sheet.cell(row=row_num, column=1, value=task.title)
        sheet.cell(row=row_num, column=2, value=task.description)
        sheet.cell(row=row_num, column=3, value=task.category)
        sheet.cell(row=row_num, column=4, value=task.due_date.strftime('%Y-%m-%d') if task.due_date else '')
        sheet.cell(row=row_num, column=5, value=task.priority)
        sheet.cell(row=row_num, column=6, value='Yes' if task.completed else 'No')
        sheet.cell(row=row_num, column=7, value=task.notes)

    for column in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=tasks.xlsx'
    workbook.save(response)
    return response


class TaskListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = TaskSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Task.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def task_list_api(request):
    tasks = Task.objects.filter(user=request.user)
    serializer = TaskSerializer(tasks, many=True)
    return Response(serializer.data)

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login') 
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})

@login_required
def task_list(request):
    query = request.GET.get('q', '')
    sort_field = request.GET.get('field', 'due_date') 
    sort_order = request.GET.get('order', 'asc')    
    sort_prefix = '' if sort_order == 'asc' else '-'
    category_filter = request.GET.get('category', '')
    priority_filter = request.GET.get('priority', '')

    now = timezone.now().date()
    near_date = now + timedelta(days=2)

    tasks = Task.objects.filter(user=request.user)
    if query:
        tasks = tasks.filter(
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(category__icontains=query)
        )

    tasks = tasks.order_by(f'{sort_prefix}{sort_field}')

    if category_filter:
        tasks = tasks.filter(category__iexact=category_filter)

    if priority_filter:
        tasks = tasks.filter(priority__iexact=priority_filter)

    paginator = Paginator(tasks, 2)
    page_number = request.GET.get('page')
    category_filter = request.GET.get('category', '')
    priority_filter = request.GET.get('priority', '')
    page_obj = paginator.get_page(page_number)

    task_count = tasks.count()
    form = TaskForm()

    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            task.save()
            return redirect('task_list')

    return render(request, 'task_list.html', {
        'form': form,
        'page_obj': page_obj,
        'query': query,
        'sort_field': sort_field,
        'sort_order': sort_order,
        'task_count': task_count,
        'category_filter': category_filter,
        'priority_filter': priority_filter,
        'now': now,
        'near_date': near_date,
    })


def task_update(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)

    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            task.save()
            return redirect('task_list')

def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)

    if request.method == 'POST':
        task.delete()
        return redirect('task_list')
    return render(request, 'task_delete.html', {'task': task})


@login_required
def task_detail_json(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    data = {
        'title': task.title,
        'description': task.description,
        'completed': task.completed,
        'due_date': task.due_date.strftime('%Y-%m-%d') if task.due_date else '',
        'category': task.category,
        'priority': task.priority,
    }
    return JsonResponse(data)